from openpyxl import Workbook, load_workbook
import string
folhas = []
alpha_col = list(string.ascii_uppercase)

def menu_inicial():
    print('______________________________________________________________________')
    print('>>>>>>>>>Estudo das funçoes essenciais da Biblioteca Openpyxl<<<<<<<<<')
    print('1 - Criar/Editar Planilha excel')
    print('2 - Consultar PLanilha excel existente')
    print('3 - Sair')


while True:
    menu_inicial()

    opcao = input('Digite uma opção:')
    if opcao == '1':
        nova_planilha = str(input('Digite o nome da planilha que deseja criar ou editar com extensão ".xlsx":'))
        extensao_arquivo = nova_planilha.endswith('xlsx')
        if extensao_arquivo == True:
            livro = Workbook()
            livro.save(nova_planilha)
            print('=>>>>>>>Planilha {} criada com sucesso!'.format(nova_planilha))
            n = int(input('Digite a quantidade de folhas trabalho que deseja criar:'))
            for indice in range(n):
                nome_folha = input('Digite o nome da folha {}:'.format(indice+1))
                folhas = livro.create_sheet(nome_folha)
                livro.save(nova_planilha)
                contador_de_folhas = livro.sheetnames

            del livro['Sheet']#apaga folha criada automaticamente por padrão.
            print('As folhas {} foram criadas com sucesso:'.format(livro.sheetnames))

            folha_escolhida = input("Agora digite o nome da folha em que deseja trabalhar:")
            for item in livro.sheetnames:
                if item == folha_escolhida:
                    folha = livro[folha_escolhida]
                    print('Folha escolhida corretamente!')
                    print('=>>>>>>>Agora Você esta trabalhando na folha: {}.'.format(folha_escolhida))
                    n_colunas = int(input("Digite o numero de colunas que deseja inserir nesta folha:"))
                    linhas = int(input('Digite o numero de filas que deseja inserir nesta folha:'))
                    conlinhas = []
                    cont_col = 0
                    cont_lin = 1

                    for x in range(linhas + 1):
                        conlinhas.append(x)
                    for i in range(linhas):
                        for n in range(n_colunas):
                            dado = input('Digite os dados da celula {}'.format(alpha_col[cont_col]) + str(cont_lin))
                            folha[str(alpha_col[cont_col]) + str(conlinhas[cont_lin])] = dado
                            cont_col = cont_col + 1
                        cont_col = 0
                        cont_lin = cont_lin + 1
                    print('=>>>>>>Dados inseridos com sucesso! Confira abaixo os dados inseridos!')
                    livro.save(nova_planilha)
                    for row in folha:
                        for cell in row:
                            print(cell.value, end=" ")
                        print()
        else:
            print('=>>>>>>Opa!Tem alguma coisa errada! Verifique se você digitou a extensão corretamente e tente novamente!')

    elif opcao == '2':
        while True:
            consulta = input('Digite o nome da planilha que deseja consultar ou "Sair" para retornar ao menu ')
            if consulta != 'sair':
                extensao_arquivo2 = consulta.endswith('xlsx')
                if extensao_arquivo2 == True:
                    try:
                        livro2 = load_workbook(consulta)
                        print('=>>>>>Planilha encontrada com sucesso!')
                        print('Esta planilha contem as seguintes folhas de trabalho:{}'.format(livro2.sheetnames))
                        escolha = input('Escolha uma folha para consulta:')
                        for item in livro2.sheetnames:
                            if item == escolha:
                                folha2 = livro2[escolha]
                                print('Folha escolhida corretamente!')
                                print('Segue os dados da folha {}'.format(escolha))
                                for row in folha2:
                                    for cell in row:
                                        print(cell.value, end=" ")
                                    print()
                                break
                            else:
                                pass
                    except FileNotFoundError:
                        print('Arquivo não enconrado!')
                else:
                    print('=>>>>>>Valor invalido! Verifique se você digitou a extensão corretamente.')
            else:
                break
        else:
                print('=>>>>>>Valor invalido! Verifique se você digitou a extensão corretamente.')

    elif opcao == '3':
        break
    else:
        print('Opcao Invalida!')
